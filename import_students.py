"""
Import students from a Kumon "Export_Student_*.xlsx" file into the local database.

Usage:
    python import_students.py path/to/Export_Student_....xlsx

Only the name (First + Last) and one parent email address are used. Everything
else in the spreadsheet (address, DOB, phone numbers, etc.) is ignored and never
touches the database. Email addresses are encrypted before being stored.

No single email column in the export covers every student, so they are tried in
order: Mother Email, Father Email, the student's own Email, then Other Email.

Safe to re-run: existing students are matched by name and their email address is
updated rather than duplicated. New names are added. It will NOT delete a student
who no longer appears in a newer export (in case they're just not in this sheet) -
use --deactivate-missing if you want that behavior.
"""
import argparse
import os
import re
import sys

import openpyxl
from dotenv import load_dotenv

sys.path.insert(0, os.path.dirname(__file__))
from app import create_app  # noqa: E402
from db import get_db  # noqa: E402
from crypto_utils import encrypt_email, normalize_email  # noqa: E402

load_dotenv()


# Checked in order - the first column with a usable address wins.
EMAIL_COLUMNS = ["Mother Email", "Father Email", "Email", "Other Email"]


def pick_email(row: dict):
    """Return (normalized_email, source_column) or (None, None)."""
    for column in EMAIL_COLUMNS:
        email = normalize_email(row.get(column))
        if email:
            return email, column
    return None, None


def build_name(row: dict) -> str:
    first = (row.get("First Name") or "").strip()
    last = (row.get("Last Name") or "").strip()
    full = f"{first} {last}".strip()
    return re.sub(r"\s+", " ", full).title()


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("xlsx_path", help="Path to the Kumon student export .xlsx file")
    parser.add_argument(
        "--deactivate-missing",
        action="store_true",
        help="Mark students not present in this file as inactive (hides them from the kiosk).",
    )
    parser.add_argument("--dry-run", action="store_true", help="Show what would happen without writing to the database.")
    args = parser.parse_args()

    wb = openpyxl.load_workbook(args.xlsx_path, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]

    imported, skipped_no_email, skipped_no_name = 0, 0, 0
    source_counts = {}
    seen_names = set()

    app = create_app()
    with app.app_context():
        db = get_db()

        for row_cells in ws.iter_rows(min_row=2, values_only=True):
            row = dict(zip(headers, row_cells))
            name = build_name(row)
            if not name:
                skipped_no_name += 1
                continue

            email, source = pick_email(row)
            if not email:
                skipped_no_email += 1
                print(f"  Skipping '{name}': no usable email address in any column.")
                continue
            source_counts[source] = source_counts.get(source, 0) + 1

            seen_names.add(name)

            if args.dry_run:
                imported += 1
                continue

            existing = db.execute("SELECT id FROM students WHERE name = %s", (name,)).fetchone()
            enc = encrypt_email(email)
            if existing:
                db.execute(
                    "UPDATE students SET email_enc = %s, active = 1 WHERE id = %s",
                    (enc, existing["id"]),
                )
            else:
                db.execute(
                    "INSERT INTO students (name, email_enc, active) VALUES (%s, %s, 1)",
                    (name, enc),
                )
            imported += 1

        if not args.dry_run:
            if args.deactivate_missing and seen_names:
                # Postgres takes the whole list as one array parameter, so there
                # is no placeholder string to build by hand.
                db.execute(
                    "UPDATE students SET active = 0 WHERE name <> ALL(%s)",
                    (list(seen_names),),
                )
            db.commit()

    print(f"\n{'Would import' if args.dry_run else 'Imported'}: {imported}")
    print(f"Skipped (no name): {skipped_no_name}")
    print(f"Skipped (no usable email address): {skipped_no_email}")
    if source_counts:
        print("Email taken from:")
        for column in EMAIL_COLUMNS:
            if source_counts.get(column):
                print(f"  {column}: {source_counts[column]}")
    if args.dry_run:
        print("\nThis was a dry run - nothing was written. Re-run without --dry-run to save.")


if __name__ == "__main__":
    main()
